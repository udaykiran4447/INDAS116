"""
Ind AS 116 – Lease Liability Working  (Streamlit)
Run:   streamlit run IndAS116_Streamlit.py
Needs: pip install streamlit openpyxl python-dateutil pandas
"""
import math, io
from datetime import date, datetime
from dateutil.relativedelta import relativedelta
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Ind AS 116 – Lease Liability",
                   page_icon="📋", layout="wide",
                   initial_sidebar_state="collapsed")

st.markdown("""<style>
[data-testid="stAppViewContainer"]{background:#F7F7F5}
.block-container{padding:1.4rem 2rem 3rem;max-width:1440px}
.ttl{background:#1A1A1A;border-radius:10px;padding:13px 22px;margin-bottom:1.1rem}
.ttl h1{color:#fff;font-size:19px;font-weight:700;margin:0}
.ttl p{color:#9CA3AF;font-size:12px;margin:2px 0 0}
.slbl{font-size:10.5px;font-weight:700;color:#6B7280;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.5rem}
.callout{background:#EFF6FF;border-left:3px solid #3B82F6;border-radius:0 6px 6px 0;padding:9px 13px;font-size:12px;color:#1E40AF;margin:4px 0 10px}
.camber{background:#FFFBEB;border-left:3px solid #F59E0B;border-radius:0 6px 6px 0;padding:9px 13px;font-size:12px;color:#92400E;margin:4px 0 10px}
label{font-size:12px !important;font-weight:600 !important;color:#374151 !important}
div[data-testid="stButton"]>button{background:#1A1A1A !important;color:#fff !important;border:none !important;border-radius:7px !important;font-weight:700 !important;font-size:13px !important;padding:9px 0 !important;width:100%}
div[data-testid="stButton"]>button:hover{background:#374151 !important}
div[data-testid="stDownloadButton"]>button{background:#1D9E75 !important;color:#fff !important;border:none !important;border-radius:7px !important;font-weight:700 !important;font-size:13px !important;padding:9px 0 !important;width:100%}
div[data-testid="stDownloadButton"]>button:hover{background:#178a64 !important}
.stTabs [data-baseweb="tab-list"]{gap:2px;background:#F3F4F6;border-radius:8px 8px 0 0;padding:4px 4px 0}
.stTabs [data-baseweb="tab"]{font-size:12.5px;font-weight:500;border-radius:6px 6px 0 0;padding:7px 18px;color:#6B7280}
.stTabs [aria-selected="true"]{background:#fff !important;color:#1A1A1A !important;border-bottom:2px solid #1A1A1A !important}
div[data-testid="stDataFrame"]{border-radius:8px;overflow:hidden}
hr{border:none;border-top:1px solid #E5E7EB;margin:.9rem 0}
</style>""", unsafe_allow_html=True)

# ── helpers ────────────────────────────────────────────────────────────────────
def add_months(dt, n): return dt + relativedelta(months=n)
def inr(n):            return f"₹ {n:,.2f}"

def parse_date(s):
    for fmt in ("%d-%m-%Y","%d/%m/%Y","%Y-%m-%d","%d-%b-%Y","%d %b %Y"):
        try: return datetime.strptime(s.strip(), fmt).date()
        except ValueError: pass
    raise ValueError(f"Cannot parse '{s}'. Use DD-MM-YYYY.")

FREQ_LABEL = {1:"Monthly",3:"Quarterly",6:"Half-Yearly",12:"Yearly"}

# ── calculation engine ─────────────────────────────────────────────────────────
def compute_schedule(from_date, to_date, lease_payment, annual_rate,
                     months_override, esc_rate, esc_every_n, pay_freq, advance):
    """
    lease_payment : cash amount paid each time (e.g. 50000 per 6-month block)
    pay_freq      : frequency in months (1=monthly, 3=quarterly, 6=half-yearly, 12=yearly)
    advance       : True = paid at START of period, False = paid at END of period

    PV discount:
      advance  → t = k * pay_freq  (first payment at t=0, DF=1)
      arrears  → t = (k+1)*pay_freq (first payment at t=pay_freq)

    Non-payment rows: cash=0, DF=blank, PV=0; interest still accrues on liability.
    Payment rows: interest on (opening – cash) if advance, else interest on opening.
    """
    mr = annual_rate / 100 / 12
    if months_override and int(months_override) > 0:
        nm = int(months_override)
    else:
        rd = relativedelta(to_date, from_date)
        nm = rd.years * 12 + rd.months
        if nm <= 0: raise ValueError("End date must be after start date.")

    n_esc      = max(1, int(esc_every_n))
    freq       = max(1, int(pay_freq))
    num_blocks = math.ceil(nm / freq)

    def block_payment(k):
        # escalation based on the month index when this payment falls
        pay_month = k * freq if advance else min((k+1)*freq - 1, nm-1)
        esc_periods = pay_month // n_esc
        return round(lease_payment * math.pow(1 + esc_rate/100, esc_periods), 2)

    cash_out = [0.0] * nm
    df_map   = {}        # month_index → discount_factor (only on payment months)
    pv       = 0.0

    for k in range(num_blocks):
        pmt     = block_payment(k)
        pay_idx = k * freq if advance else min((k+1)*freq - 1, nm-1)
        t       = pay_idx  if advance else pay_idx + 1   # time in months for discounting
        if pay_idx < nm:
            cash_out[pay_idx] = pmt
            df                = 1 / math.pow(1 + mr, t)
            df_map[pay_idx]   = df
            pv               += pmt * df

    rou_depr = pv / nm
    rows, open_liab = [], pv

    for i in range(nm):
        cash  = cash_out[i]
        df    = df_map.get(i, None)          # None → blank on non-payment rows
        pv_col = round(cash * df, 2) if (df is not None and cash > 0) else 0.0

        if advance and cash > 0:
            # advance payment: deduct cash first, then interest on remainder
            after    = open_liab - cash
            interest = after * mr
            cl       = after + interest
        else:
            # arrears or non-payment month: interest on full opening
            interest = open_liab * mr
            cl       = open_liab + interest - cash

        cl  = 0.0 if abs(cl)  < 0.005 else cl
        rou = pv - rou_depr * (i + 1)
        rou = 0.0 if abs(rou) < 0.005 else rou

        rows.append({
            "sno":    i + 1,
            "date":   add_months(from_date, i + 1),
            "mp":     cash_out[i],
            "df":     df,
            "pv_col": pv_col,
            "cash":   cash,
            "open":   round(open_liab, 2),
            "int":    round(interest, 2),
            "close":  round(cl, 2),
            "depr":   round(rou_depr, 2),
            "rou":    round(rou, 2),
        })
        open_liab = cl

    return rows, pv, mr, nm, rou_depr


# ── Excel builder ──────────────────────────────────────────────────────────────
def build_excel(schedule, info):
    """
    ARREARS → 11 cols A–K:
      A:S.No  B:Month  C:Lease Payment  D:Discount Factor  E:Present Value
      F:Opening Liability  G:Interest Expense  H:Lease Payment(cash)
      I:Closing Liability  J:Depreciation  K:ROU Asset

    ADVANCE → 12 cols A–L (extra col D = Advance Cash Paid):
      A:S.No  B:Month  C:Lease Payment  D:Advance Cash Paid  E:Discount Factor
      F:Present Value  G:Opening Liability  H:Interest Expense  I:Lease Payment(cash)
      J:Closing Liability  K:Depreciation  L:ROU Asset

    Key formula rules (no Python operators in Excel strings):
      - Escalation : =$B$10*(1+$B$13/100)^INT((A{r}-1)/$B$14)
      - Discount t  : hardcoded integer exponent derived from df value
      - C non-pay   : hardcoded 0 (arrears, non-payment row)
      - D/E non-pay : empty string (blank cell)
    """
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = info["name"][:31]
    ws.sheet_view.showGridLines = False

    adv   = info["advance"]
    NM    = info["num_months"]
    freq  = info["pay_freq"]
    ar    = info["annual_rate"]
    mr    = ar / 100 / 12
    NF    = "#,##0.00"
    NCOLS = 12 if adv else 11

    thin = Side(border_style="thin", color="D1D5DB")
    def bdr(): return Border(left=thin, right=thin, top=thin, bottom=thin)

    def W(row, col, val, bold=False, align="left", fmt=None,
          bg=None, fg="1A1A1A", size=10, wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name="Calibri", bold=bold, size=size, color=fg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if bg:  c.fill          = PatternFill("solid", fgColor=bg)
        if fmt: c.number_format = fmt
        return c

    def MW(r, c1, c2, val, **kw):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        W(r, c1, val, **kw)

    def t_from_df(df_val):
        """Recover integer t from df = 1/(1+mr)^t"""
        if df_val is None or df_val <= 0:
            return None
        if abs(df_val - 1.0) < 1e-9:
            return 0
        return int(round(-math.log(df_val) / math.log(1 + mr)))

    # ── Rows 1–2: Title & lease name ─────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    MW(1, 1, NCOLS, "IND AS 116 – LEASE LIABILITY WORKING",
       bold=True, align="center", size=14, bg="1A1A1A", fg="FFFFFF")
    ws.row_dimensions[2].height = 22
    MW(2, 1, NCOLS, info["name"],
       bold=True, align="center", size=12, bg="1D9E75", fg="FFFFFF")
    ws.row_dimensions[3].height = 6

    # ── Info block rows 4–14 ─────────────────────────────────────────────────
    freq_label = FREQ_LABEL.get(freq, f"Every {freq} months")
    adv_label  = f"{'Advance' if adv else 'Arrears'} – {freq_label}"
    meta = [
        (4,  "Lease Term (Years)",                    round(NM/12, 4),         "0.00",  "0000FF"),
        (5,  "Beginning From",                        info["from_date"].strftime("%d-%b-%Y"), None, "0000FF"),
        (6,  "Ending To",                             info["to_date"].strftime("%d-%b-%Y"),   None, "0000FF"),
        (7,  "No. of Months",                         NM,                      "0",     "0000FF"),
        (8,  "Annual Discount Rate",                  ar / 100,                "0.00%", "0000FF"),
        (9,  "Monthly Discount Rate",                 "=B8/12",                "0.00%", "1A1A1A"),
        (10, "Lease Payment (Rs.)",                   info["lease_payment"],   NF,      "0000FF"),
        (11, "Payment Frequency",                     freq_label,              None,    "0000FF"),
        (12, "Payment Mode",                          adv_label,               None,    "0000FF"),
        (13, "Escalation Rate (%)",                   info["esc_rate"],         "0.00",  "0000FF"),
        (14, "Escalation Frequency (every N months)", info["esc_every_n"],      "0",     "0000FF"),
    ]
    for (row_n, lbl, val, fmt, fg) in meta:
        ws.row_dimensions[row_n].height = 15
        W(row_n, 1, lbl, bold=True, fg="374151", size=9)
        W(row_n, 2, val, fmt=fmt,   fg=fg,       size=9)

    PV_ROW = 15
    ws.row_dimensions[PV_ROW].height = 16
    W(PV_ROW, 1, "PV of Lease Payments – Initial Lease Liability",
      bold=True, fg="374151", size=9)
    W(PV_ROW, 2, round(info["pv"], 2), fmt=NF, bold=True, size=10, fg="1D9E75")
    ws.row_dimensions[16].height = 8
    ws.row_dimensions[17].height = 8

    # ── Header row 18, data starts row 19 ────────────────────────────────────
    HDR_ROW    = 18
    DATA_START = 19
    PV_REF     = f"$B${PV_ROW}"    # =$B$15
    NM_REF     = "$B$7"            # No. of months

    ws.row_dimensions[HDR_ROW].height = 44
    if not adv:
        HDR_LABELS = [
            "S.No", "Month",
            "Lease\nPayment\n(Rs.)", "Discount\nFactor", "Present\nValue\n(Rs.)",
            "Opening\nLiability\n(Rs.)", "Interest\nExpense\n(Rs.)",
            "Lease\nPayment\n(Rs.)", "Closing\nLiability\n(Rs.)",
            "Depreciation\n(Rs.)", "ROU Asset\n(Rs.)",
        ]
        COL_WIDTHS = [6, 13, 15, 13, 15, 18, 15, 15, 18, 15, 15]
    else:
        HDR_LABELS = [
            "S.No", "Month",
            "Lease\nPayment\n(Rs.)", "Advance\nCash Paid\n(Rs.)",
            "Discount\nFactor", "Present\nValue\n(Rs.)",
            "Opening\nLiability\n(Rs.)", "Interest\nExpense\n(Rs.)",
            "Lease\nPayment\n(Rs.)", "Closing\nLiability\n(Rs.)",
            "Depreciation\n(Rs.)", "ROU Asset\n(Rs.)",
        ]
        COL_WIDTHS = [6, 13, 15, 16, 13, 15, 18, 15, 15, 18, 15, 15]

    for ci, hdr in enumerate(HDR_LABELS, 1):
        c = W(HDR_ROW, ci, hdr, bold=True, align="center",
               bg="1A1A1A", fg="FFFFFF", size=9, wrap=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr()
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, row in enumerate(schedule):
        r      = DATA_START + ri
        bg     = "F3F4F6" if ri % 2 else None
        is_pay = row["cash"] > 0
        df_val = row["df"]           # None on non-payment rows
        t      = t_from_df(df_val)  # integer discount exponent (None if non-payment)
        ws.row_dimensions[r].height = 15

        def wr(col, val, fmt=None, align="right", _r=r, _bg=bg):
            c2 = W(_r, col, val, fmt=fmt, align=align, bg=_bg, size=9)
            c2.border = bdr()
            return c2

        # ── ESCALATION formula helper (valid Excel, no Python operators) ──────
        # =$B$10*(1+$B$13/100)^INT((A{r}-1)/$B$14)
        # Uses cell A{r} which contains the S.No (row number 1-based)
        esc_formula = f"=$B$10*(1+$B$13/100)^INT((A{r}-1)/$B$14)"

        if not adv:
            # ══════════════════════════════════════════════════════════════════
            # ARREARS – 11 columns A to K
            # ══════════════════════════════════════════════════════════════════
            # A: S.No (hardcoded)
            wr(1, ri + 1, fmt="0", align="center")

            # B: Month (hardcoded date string)
            wr(2, row["date"].strftime("%d-%b-%Y"), align="center")

            # C: Lease Payment
            #    payment row  → escalation formula (driven by B10, B13, B14)
            #    non-payment  → hardcoded 0
            if is_pay:
                wr(3, esc_formula, fmt=NF)
            else:
                wr(3, 0, fmt=NF)

            # D: Discount Factor
            #    payment row  → =1/(1+$B$9)^t  where t is integer
            #    non-payment  → blank
            if is_pay and t is not None:
                wr(4, f"=1/(1+$B$9)^{t}", fmt="0.00000000")
            else:
                wr(4, None, fmt="0.00000000")

            # E: Present Value = C * D
            #    payment row  → =C{r}*D{r}
            #    non-payment  → blank
            if is_pay:
                wr(5, f"=C{r}*D{r}", fmt=NF)
            else:
                wr(5, None, fmt=NF)

            # F: Opening Liability
            #    row 1 → =$B$15 (PV)
            #    rest  → =I{r-1} (previous closing)
            wr(6, f"={PV_REF}" if ri == 0 else f"=I{r-1}", fmt=NF)

            # G: Interest Expense = F * $B$9  (interest on full opening, arrears)
            wr(7, f"=F{r}*$B$9", fmt=NF)

            # H: Lease Payment (cash out) = C  (mirrors C column)
            wr(8, f"=C{r}", fmt=NF)

            # I: Closing Liability = F + G – H
            wr(9, f"=F{r}+G{r}-H{r}", fmt=NF)

            # J: Depreciation = PV / No. of months
            wr(10, f"={PV_REF}/{NM_REF}", fmt=NF)

            # K: ROU Asset = PV – J * S.No
            wr(11, f"={PV_REF}-J{r}*A{r}", fmt=NF)

        else:
            # ══════════════════════════════════════════════════════════════════
            # ADVANCE – 12 columns A to L
            # ══════════════════════════════════════════════════════════════════
            # A: S.No
            wr(1, ri + 1, fmt="0", align="center")

            # B: Month
            wr(2, row["date"].strftime("%d-%b-%Y"), align="center")

            # C: Monthly lease payment equivalent (escalation formula every row)
            wr(3, esc_formula, fmt=NF)

            # D: Advance Cash Paid
            #    payment row  → =SUM(C{r}:C{block_end})  (sum of freq months)
            #    non-payment  → hardcoded 0
            if is_pay:
                block_end_r = min(r + freq - 1, DATA_START + NM - 1)
                wr(4, f"=SUM(C{r}:C{block_end_r})", fmt=NF)
            else:
                wr(4, 0, fmt=NF)

            # E: Discount Factor
            #    payment row  → =1/(1+$B$9)^t
            #    non-payment  → blank
            if is_pay and t is not None:
                wr(5, f"=1/(1+$B$9)^{t}", fmt="0.00000000")
            else:
                wr(5, None, fmt="0.00000000")

            # F: Present Value = C * E on payment rows, blank otherwise
            if is_pay:
                wr(6, f"=C{r}*E{r}", fmt=NF)
            else:
                wr(6, None, fmt=NF)

            # G: Opening Liability
            wr(7, f"={PV_REF}" if ri == 0 else f"=J{r-1}", fmt=NF)

            # H: Interest Expense = (G – D) * $B$9
            #    (advance: interest on opening after deducting cash paid)
            wr(8, f"=(G{r}-D{r})*$B$9", fmt=NF)

            # I: Lease Payment (cash out) = D  (mirrors D column)
            wr(9, f"=D{r}", fmt=NF)

            # J: Closing Liability = G – D + H
            wr(10, f"=G{r}-D{r}+H{r}", fmt=NF)

            # K: Depreciation = PV / No. of months
            wr(11, f"={PV_REF}/{NM_REF}", fmt=NF)

            # L: ROU Asset = PV – K * S.No
            wr(12, f"={PV_REF}-K{r}*A{r}", fmt=NF)

    # ── Totals row ────────────────────────────────────────────────────────────
    TOT = DATA_START + NM
    ws.row_dimensions[TOT].height = 17
    # Col A: label (not merged → safe)
    c = W(TOT, 1, "TOTALS", bold=True, align="right", bg="374151", fg="FFFFFF", size=9)
    c.border = bdr()
    # Col B: style only (no value — avoids MergedCell read-only error)
    ws.cell(TOT, 2).fill   = PatternFill("solid", fgColor="374151")
    ws.cell(TOT, 2).border = bdr()
    # Cols 3+: SUM formulas for relevant cols, dash for others
    if not adv:
        sum_cols   = {3, 5, 7, 8}       # C=pmt, E=PV, G=interest, H=cash
        total_cols = 11
    else:
        sum_cols   = {3, 4, 6, 8, 9}    # C=pmt, D=adv cash, F=PV, H=interest, I=cash
        total_cols = 12
    for ci in range(3, total_cols + 1):
        col = get_column_letter(ci)
        if ci in sum_cols:
            c = W(TOT, ci, f"=SUM({col}{DATA_START}:{col}{TOT-1})",
                  fmt=NF, bold=True, bg="374151", fg="FFFFFF", size=9, align="right")
        else:
            c = W(TOT, ci, "—", bold=True, bg="374151", fg="9CA3AF", size=9, align="center")
        c.border = bdr()

    # ── Year-1 journal summary ────────────────────────────────────────────────
    SUM_R = TOT + 3
    ws.row_dimensions[SUM_R].height = 18
    MW(SUM_R, 1, NCOLS, "YEAR 1 – JOURNAL ENTRY SUMMARY",
       bold=True, align="center", bg="374151", fg="FFFFFF", size=10)
    yr1_end  = DATA_START + min(12, NM) - 1
    depr_col = "J" if not adv else "K"
    int_col  = "G" if not adv else "H"
    cash_col = "H" if not adv else "D"
    jrows = [
        ("Depreciation A/c",                "Dr", f"=SUM({depr_col}{DATA_START}:{depr_col}{yr1_end})", ""),
        ("   To ROU Asset A/c",             "Cr", "", f"=SUM({depr_col}{DATA_START}:{depr_col}{yr1_end})"),
        ("", "", "", ""),
        ("Interest on Lease Liability A/c", "Dr", f"=SUM({int_col}{DATA_START}:{int_col}{yr1_end})", ""),
        ("   To Lease Liability A/c",       "Cr", "", f"=SUM({int_col}{DATA_START}:{int_col}{yr1_end})"),
        ("", "", "", ""),
        ("Lease Liability A/c",             "Dr", f"=SUM({cash_col}{DATA_START}:{cash_col}{yr1_end})", ""),
        ("   To Bank / Cash A/c",           "Cr", "", f"=SUM({cash_col}{DATA_START}:{cash_col}{yr1_end})"),
    ]
    for ji, (part, dc2, dr, cr) in enumerate(jrows):
        row_n = SUM_R + 2 + ji
        ws.row_dimensions[row_n].height = 14
        W(row_n, 1, part, size=9, fg="111827")
        W(row_n, 2, dc2,  size=9, fg="6B7280", bold=True, align="center")
        if dr: W(row_n, 3, dr, fmt=NF, align="right", size=9)
        if cr: W(row_n, 4, cr, fmt=NF, align="right", size=9)

    ws.freeze_panes = f"A{DATA_START}"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""<div class="ttl">
  <h1>📋 Ind AS 116 – Lease Liability Working</h1>
  <p>Amortisation schedule · ROU asset · Journal entries · Excel with live formulas</p>
</div>""", unsafe_allow_html=True)

# ── Section 1: Lease Details ───────────────────────────────────────────────────
st.markdown('<div class="slbl">Lease Details</div>', unsafe_allow_html=True)
c1,c2,c3 = st.columns([2,1,1])
with c1: lease_name = st.text_input("Lease Name", placeholder="e.g. Gachibowli Office – 3rd Floor")
with c2: lease_from = st.text_input("Lease Start Date", placeholder="DD-MM-YYYY")
with c3: lease_to   = st.text_input("Lease End Date",   placeholder="DD-MM-YYYY")
c4,c5,c6 = st.columns(3)
with c4: annual_rate     = st.number_input("Annual Discount Rate (%)", min_value=0.01,max_value=50.0,value=11.0,step=0.01,format="%.2f")
with c5: months_override = st.number_input("No. of Months  (0 = auto from dates)", min_value=0,value=0,step=1)
with c6: st.markdown("")

st.markdown("---")

# ── Section 2: Payment Details ─────────────────────────────────────────────────
st.markdown('<div class="slbl">Payment Details</div>', unsafe_allow_html=True)

pa,pb,pc,pd_ = st.columns([1,1,1,1])
with pa:
    lease_payment = st.number_input(
        "Lease Payment (₹)",
        min_value=0.0, value=0.0, step=1000.0, format="%.2f",
        help="The actual cash amount paid each time (e.g. 50000 per 6-month block, not the monthly equivalent).")
with pb:
    pay_freq = st.number_input(
        "Payment Frequency (months)",
        min_value=1, max_value=120, value=1, step=1,
        help="How often the payment is made.\n1 = Monthly | 3 = Quarterly | 6 = Half-Yearly | 12 = Yearly\nYou can enter any number of months.")
with pc:
    freq_hint = FREQ_LABEL.get(pay_freq, f"Every {pay_freq} months")
    st.markdown(f"<br><span style='font-size:13px;color:#1D9E75;font-weight:600'>{freq_hint}</span>",
                unsafe_allow_html=True)
with pd_:
    advance_sel = st.selectbox("Paid in Advance?",
                               ["No – Arrears (end of period)", "Yes – Advance (start of period)"])
    advance = advance_sel.startswith("Yes")

# Callout
if lease_payment > 0:
    if advance:
        st.markdown(
            f'<div class="camber">⚠️ <b>Advance – {freq_hint}:</b> '
            f'₹{lease_payment:,.2f} paid at the <b>start</b> of every {pay_freq}-month block. '
            f'First payment discount factor = <b>1.000</b> (t=0). '
            f'Excel will have 12 columns including "Advance Cash Paid".</div>',
            unsafe_allow_html=True)
    else:
        st.markdown(
            f'<div class="callout">ℹ️ <b>Arrears – {freq_hint}:</b> '
            f'₹{lease_payment:,.2f} paid at the <b>end</b> of every {pay_freq}-month block. '
            f'Excel will have 11 columns.</div>',
            unsafe_allow_html=True)

st.markdown("---")

# ── Section 3: Escalation ──────────────────────────────────────────────────────
st.markdown('<div class="slbl">Escalation Settings</div>', unsafe_allow_html=True)
ec1,ec2 = st.columns(2)
with ec1:
    esc_rate = st.number_input("Escalation Rate (%)", min_value=0.0,max_value=100.0,
                               value=0.0,step=0.01,format="%.2f",
                               help="% increase at each escalation step. 0 = no escalation.")
with ec2:
    esc_every_n = st.number_input("Escalation Every N Months", min_value=1,max_value=360,
                                  value=12,step=1,
                                  help="Payment steps up every N months. 12=annual | 6=semi-annual | 3=quarterly | 1=monthly")
if esc_rate > 0 and lease_payment > 0:
    after = round(lease_payment*(1+esc_rate/100),2)
    st.markdown(f'<div class="callout">💡 Payment increases by <b>{esc_rate:.2f}%</b> every '
                f'<b>{esc_every_n} month(s)</b>: ₹{lease_payment:,.2f} → ₹{after:,.2f}.</div>',
                unsafe_allow_html=True)

st.markdown("")
bcol,_ = st.columns([1,4])
with bcol:
    do_calc = st.button("⚡  Calculate Schedule", use_container_width=True)

# ── Calculation ────────────────────────────────────────────────────────────────
if do_calc:
    errs=[]
    if not lease_name.strip(): errs.append("Lease name is required.")
    if lease_payment <= 0:     errs.append("Lease payment must be > 0.")
    from_date=to_date=None
    if lease_from.strip():
        try:    from_date=parse_date(lease_from)
        except ValueError as e: errs.append(str(e))
    else: from_date=date.today()
    if lease_to.strip():
        try:    to_date=parse_date(lease_to)
        except ValueError as e: errs.append(str(e))
    if months_override==0 and to_date is None:
        errs.append("Enter lease end date or number of months.")
    for e in errs: st.error(e)
    if errs: st.stop()

    eff_to = to_date if to_date else add_months(from_date, int(months_override))
    try:
        sched,pv,mr,nm,rou_d = compute_schedule(
            from_date,eff_to,lease_payment,annual_rate,
            months_override,esc_rate,esc_every_n,pay_freq,advance)
    except Exception as e:
        st.error(f"Calculation error: {e}"); st.stop()

    freq_label = FREQ_LABEL.get(pay_freq, f"Every {pay_freq} months")
    st.session_state.update({
        "sched":sched,"pv":pv,"mr":mr,"nm":nm,"rou_d":rou_d,
        "info":{
            "name":lease_name.strip() or "Lease",
            "from_date":from_date,"to_date":eff_to,
            "num_months":nm,"annual_rate":annual_rate,"monthly_rate":mr,
            "pv":pv,"lease_payment":lease_payment,
            "pay_freq":pay_freq,"advance":advance,
            "freq_label":freq_label,
            "adv_label":f"{'Advance' if advance else 'Arrears'} – {freq_label}",
            "esc_rate":esc_rate,"esc_every_n":esc_every_n,"rou_depr":rou_d,
        }
    })

# ── Results ────────────────────────────────────────────────────────────────────
if "sched" in st.session_state:
    sched=st.session_state["sched"]; pv=st.session_state["pv"]
    mr=st.session_state["mr"];       nm=st.session_state["nm"]
    rou_d=st.session_state["rou_d"]; info=st.session_state["info"]
    adv=info["advance"];              freq=info["pay_freq"]

    st.markdown("---")
    m1,m2,m3,m4,m5,m6=st.columns(6)
    m1.metric("Lease Term",              f"{nm} months")
    m2.metric("Annual Discount Rate",    f"{info['annual_rate']:.2f}%")
    m3.metric("Monthly Discount Rate",   f"{mr*100:.4f}%")
    m4.metric("Initial Lease Liability", inr(pv))
    m5.metric("Monthly Depreciation",    inr(rou_d))
    m6.metric("Payment Mode",            info["adv_label"])
    st.markdown("")

    tab1,tab2,tab3,tab4=st.tabs([
        "  📊 Amortisation Schedule  ",
        "  💸 Cash Payment Schedule  ",
        "  📖 Journal Entries (Yr 1)  ",
        "  📋 Payment Step Summary  ",
    ])

    # ── Tab 1 ─────────────────────────────────────────────────────────────────
    with tab1:
        rows_d=[]
        for r in sched:
            row_d={"S.No":r["sno"],"Month":r["date"].strftime("%d-%b-%Y")}
            if adv:
                row_d["Advance Cash Paid (₹)"]=f"{r['cash']:,.2f}" if r["cash"] else "—"
            else:
                row_d["Lease Payment (₹)"]=f"{r['cash']:,.2f}" if r["cash"] else "—"
            row_d["Discount Factor"]=f"{r['df']:.8f}" if r["df"] is not None else "—"
            row_d["Present Value (₹)"]=f"{r['pv_col']:,.2f}" if r["pv_col"] else "—"
            row_d.update({
                "Opening Liability (₹)":f"{r['open']:,.2f}",
                "Interest Expense (₹)": f"{r['int']:,.2f}",
                "Closing Liability (₹)":f"{r['close']:,.2f}",
                "Depreciation (₹)":     f"{r['depr']:,.2f}",
                "ROU Asset (₹)":        f"{r['rou']:,.2f}",
            })
            rows_d.append(row_d)
        st.dataframe(pd.DataFrame(rows_d),use_container_width=True,hide_index=True,height=430)
        pay_rows=[r for r in sched if r["cash"]>0]
        st.caption(
            f"Rows: {nm}  ·  Payment events: {len(pay_rows)}  ·  "
            f"Total cash paid: {inr(sum(r['cash'] for r in sched))}  ·  "
            f"Total interest: {inr(sum(r['int'] for r in sched))}  ·  "
            f"Total depreciation: {inr(sum(r['depr'] for r in sched))}")

    # ── Tab 2 ─────────────────────────────────────────────────────────────────
    with tab2:
        cash_rows=[r for r in sched if r["cash"]>0]
        df_cash=[]
        for r in cash_rows:
            blk_start=r["sno"]
            blk_end  =min(r["sno"]+freq-1,nm) if adv else r["sno"]
            df_cash.append({
                "S.No":r["sno"],
                "Payment Date":r["date"].strftime("%d-%b-%Y"),
                "Period":f"Month {blk_start}{'–'+str(blk_end) if blk_end!=blk_start else ''}",
                "Cash Paid (₹)":f"{r['cash']:,.2f}",
                "Discount Factor":f"{r['df']:.8f}" if r["df"] is not None else "—",
                "Present Value (₹)":f"{r['pv_col']:,.2f}",
            })
        st.markdown("**Actual cash outflows with discount factors**")
        st.dataframe(pd.DataFrame(df_cash),use_container_width=True,hide_index=True,
                     height=min(450,len(df_cash)*38+50))
        st.caption(f"Total payment events: {len(cash_rows)}  ·  "
                   f"Total: {inr(sum(r['cash'] for r in cash_rows))}")

    # ── Tab 3 ─────────────────────────────────────────────────────────────────
    with tab3:
        yr1=sched[:min(12,nm)]
        yr1_d=sum(r["depr"] for r in yr1); yr1_i=sum(r["int"] for r in yr1)
        yr1_p=sum(r["cash"] for r in yr1)
        st.markdown("**Initial Recognition (Day 1)**")
        st.dataframe(pd.DataFrame([
            {"Particulars":"ROU Asset A/c","Dr (₹)":f"{pv:,.2f}","Cr (₹)":"—"},
            {"Particulars":"   To Lease Liability A/c","Dr (₹)":"—","Cr (₹)":f"{pv:,.2f}"},
        ]),hide_index=True,use_container_width=True,height=100)
        st.markdown("**Monthly entries – Year 1**")
        jrows=[]
        for r in yr1:
            d=r["date"].strftime("%d-%b-%Y")
            if r["cash"]>0 and adv:
                jrows+=[
                    {"Date":d,"Particulars":"Lease Liability A/c  Dr","Dr (₹)":f"{r['cash']:,.2f}","Cr (₹)":"—","Narration":"Advance cash paid"},
                    {"Date":"","Particulars":"   To Bank / Cash A/c","Dr (₹)":"—","Cr (₹)":f"{r['cash']:,.2f}","Narration":""},
                ]
            jrows+=[
                {"Date":d,"Particulars":"Interest on Lease Liability A/c  Dr","Dr (₹)":f"{r['int']:,.2f}","Cr (₹)":"—","Narration":"Interest accrual"},
                {"Date":"","Particulars":"   To Lease Liability A/c","Dr (₹)":"—","Cr (₹)":f"{r['int']:,.2f}","Narration":""},
                {"Date":d,"Particulars":"Depreciation A/c  Dr","Dr (₹)":f"{r['depr']:,.2f}","Cr (₹)":"—","Narration":"ROU depreciation"},
                {"Date":"","Particulars":"   To ROU Asset A/c","Dr (₹)":"—","Cr (₹)":f"{r['depr']:,.2f}","Narration":""},
                {"Date":"·","Particulars":"·","Dr (₹)":"","Cr (₹)":"","Narration":""},
            ]
            if not adv and r["cash"]>0:
                jrows.insert(-1,{"Date":d,"Particulars":"Lease Liability A/c  Dr","Dr (₹)":f"{r['cash']:,.2f}","Cr (₹)":"—","Narration":"Lease payment"})
                jrows.insert(-1,{"Date":"","Particulars":"   To Bank / Cash A/c","Dr (₹)":"—","Cr (₹)":f"{r['cash']:,.2f}","Narration":""})
        st.dataframe(pd.DataFrame(jrows),hide_index=True,use_container_width=True,height=430)
        last_yr1=sched[min(11,nm-1)]
        st.markdown("**Year 1 Summary**")
        st.dataframe(pd.DataFrame([
            {"Particulars":"Total Depreciation (Yr 1)",    "Amount (₹)":f"{yr1_d:,.2f}"},
            {"Particulars":"Total Interest Expense (Yr 1)","Amount (₹)":f"{yr1_i:,.2f}"},
            {"Particulars":"Total Cash Paid (Yr 1)",       "Amount (₹)":f"{yr1_p:,.2f}"},
            {"Particulars":"Closing Lease Liability",      "Amount (₹)":f"{last_yr1['close']:,.2f}"},
            {"Particulars":"Closing ROU Asset",            "Amount (₹)":f"{last_yr1['rou']:,.2f}"},
        ]),hide_index=True,use_container_width=True,height=215)

    # ── Tab 4 ─────────────────────────────────────────────────────────────────
    with tab4:
        pay_rows=[r for r in sched if r["cash"]>0]
        steps=[]
        for r in pay_rows:
            steps.append({
                "S.No":r["sno"],
                "Payment Date":r["date"].strftime("%d-%b-%Y"),
                "Cash Amount (₹)":f"{r['cash']:,.2f}",
                "Discount Factor":f"{r['df']:.8f}" if r["df"] is not None else "—",
                "PV (₹)":f"{r['pv_col']:,.2f}",
            })
        st.markdown("**All payment events across the lease term**")
        st.dataframe(pd.DataFrame(steps),hide_index=True,use_container_width=True)
        st.caption(f"Frequency: {info['adv_label']}  ·  "
                   f"Total events: {len(steps)}  ·  "
                   f"Total cash: {inr(sum(r['cash'] for r in pay_rows))}")

    # ── Download ───────────────────────────────────────────────────────────────
    st.markdown("---")
    dcol,icol=st.columns([1,3])
    with dcol:
        xl=build_excel(sched,info)
        fname=f"IndAS116_{info['name'].replace(' ','_')}.xlsx"
        st.download_button(label="⬇️  Download Excel",data=xl,file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    with icol:
        cols_n="11 columns" if not adv else "12 columns (incl. Advance Cash Paid)"
        st.caption(f"Excel: {cols_n}  ·  all cells use live Excel formulas  ·  "
                   "blue=inputs · formula cells=auto-calculated · green=PV  ·  "
                   "non-payment rows show blank DF/PV as per Ind AS 116")
